import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

workbook = Workbook()
workbook.remove(workbook.active)
script = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(script, 'uSch_JSON')

startH = 8
startM = 0
endH = 22
endM = 0

def comparer(week):
    #sets time slots in spreadsheet
    newSheet = workbook.create_sheet(title = week)
    hour = startH
    minute = startM
    x = 1
    while hour < endH:
        x+=1
        timeStr = f"{hour:02}:{minute:02}"
        newSheet.cell(column = 1, row = x).value = timeStr
        minute+= 30
        if minute==60:
            hour+=1
            minute-=60

    col = 1
    for file_path in os.listdir(path):
        col+=1
        newSheet.cell(row = 1, column = col).value = file_path[:-5]
        Time = []
        for x in range(28):
            Time.append('.')
        with open(path+'\\'+file_path) as f:
            json_data = json.load(f)
            for x in range(len(json_data['courses'])):
                for y in range(len(json_data['courses'][x]['sections'][0]['components'])):
                    if(json_data['courses'][x]['sections'][0]['components'][y]['day'] == week):
                        start = json_data['courses'][x]['sections'][0]['components'][y]['start_time']
                        
                        timeStart = ((int(start[:2])-8)*2) + (int(start[3:])//20)
                        end = json_data['courses'][x]['sections'][0]['components'][y]['end_time']
                        timeEnd = ((int(end[:2])-8)*2) + (int(end[3:])//20)
                        for z in range(timeStart, timeEnd):
                            
                            Time[z] = 'X'
                    
        for x in range(28):
            color = colors[Time[x]]
            newSheet.cell(column = os.listdir(path).index(file_path)+2, row = x+2).fill = PatternFill(start_color=(color), end_color=(color), fill_type='solid') 
    return Time
    

#main
colors = {'.': 'AAEEB0', 'X':'EEAAAA'}
days = ['MO','TU','WE','TH','FR','SA','SU']
for day in days:
    comparer(day)
        
workbook.save('schedule.xlsx')
